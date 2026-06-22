"""Global DCR, ORR, and durable clinical benefit (DB) analysis pipeline.

Computes response-rate summaries and pairwise Fisher exact tests (with
Benjamini–Hochberg FDR correction) for three DDA stratification schemes:

  1. On-label utility — LEVEL tiers with on/off-label splits and standard-of-care
     chemotherapy reference (SC).
  2. Compound ranking — per-patient dense LEVEL rank (1–6, pooled 7–10) with SC.
  3. SHIVA tiers — Low / Intermediate / High by LEVEL.

Endpoints
---------
  ORR  — objective response rate (CR + PR)
  DCR  — disease control rate (CR + PR + SD)
  DB   — durable clinical benefit (TimeToDouble > 42 days)

Response classifications are derived from mRECIST thresholds applied to
BestResponse and BestAvgResponse. Cochran–Armitage trend tests are computed
over ordered tier groups (excluding pooled reference arms).

Usage
-----
Run from the repository root:

python src/DCR_ORR_DB_PDX.py [--source PATH] [--output-dir DIR]
                                       [--analyses {all,on_label,ranking,shiva} ...]

Outputs are written under ``<output-dir>/{on_label,ranking,shiva}/``.
A ``run_metadata.json`` file records input checksums for reproducibility.
Random seed: not applicable (deterministic given fixed inputs).
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
from datetime import datetime
from itertools import combinations
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from scipy import stats
from scipy.stats import fisher_exact
from statsmodels.stats.multitest import multipletests

LOGGER = logging.getLogger("DCR_ORR_DB_PDX")

# ---------------------------------------------------------------------------
# Repository-relative paths (no hard-coded absolute paths)
# ---------------------------------------------------------------------------

_SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = _SCRIPT_DIR.parent
DATA_DIR = REPO_ROOT / "data"

DEFAULT_SOURCE = DATA_DIR / "pdx_curve_metrics_single_treatments_dda_scores.xlsx"
DEFAULT_CHEMO_CSV = DATA_DIR / "chemo_data_2024_12_06.csv"
DEFAULT_CHEMO_XLSX = DATA_DIR / "chemo_data_2024_12_06_exc.xlsx"
DEFAULT_OUTPUT_ROOT = _SCRIPT_DIR / "DCR_ORR_DB"

RUN_DATE = datetime.now().strftime("%Y_%m_%d")

TREAT_TYPE = "single"
DCB_DAYS_THRESHOLD = 42
RANK_POOL_FROM = 7
RANK_POOLED_LABEL = "7-10"

RESPONSE_CLASSIFICATIONS = [
    "Response_Best_Response",
    "Response_BestAvgResponse",
]
RESPONSE_CLASS_LABEL = {
    "Response_Best_Response": "Best_Response",
    "Response_BestAvgResponse": "BestAvgResponse",
}

ENDPOINTS_ORR_DCR = ("DCR", "ORR")
ENDPOINT_DB = "DB"

COCHRAN_EXCLUDE_GROUPS = frozenset({"all", "SC"})

COCHRAN_GROUP_ORDER = {
    "on_label": ["Low", "Int", "High", "High On", "High Off", "High Off+Exp"],
    "ranking": ["1", "2", "3", "4", "5", "6", "7-10"],
    "shiva": ["Low", "Intermediate", "High"],
}

# Ranking pairwise comparisons: rank 1 vs each lower tier and SC reference.
RANKING_FISHER_PAIRS: Tuple[Tuple[str, str], ...] = tuple(
    ("1", g) for g in ("2", "3", "4", "5", "6", RANK_POOLED_LABEL, "SC")
)

SCI_NUMBER_FORMAT = "0.00E+00"
ODDS_RATIO_NUMBER_FORMAT = "0.000"

SOURCE_XLSX: Path = DEFAULT_SOURCE
CHEMO_CSV: Path = DEFAULT_CHEMO_CSV
CHEMO_XLSX: Path = DEFAULT_CHEMO_XLSX
OUTPUT_ROOT: Path = DEFAULT_OUTPUT_ROOT

ANALYSIS_METHODS = ("on_label", "ranking", "shiva")

# ---------------------------------------------------------------------------
# mRECIST response categories and durable-benefit endpoint
# ---------------------------------------------------------------------------


def categorize_best_avg_response(x: float) -> str:
    """Map BestAvgResponse (percent change) to mRECIST category."""
    if x < -40:
        return "CR"
    if x < -20:
        return "PR"
    if x < 30:
        return "SD"
    return "PD"


def categorize_best_response(x: float) -> str:
    """Map BestResponse (percent change) to mRECIST category."""
    if x < -95:
        return "CR"
    if x < -50:
        return "PR"
    if x < 35:
        return "SD"
    return "PD"


def assign_mrecist_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Derive mRECIST response columns from continuous tumour-volume metrics."""
    out = df.copy()
    out["Response_BestAvgResponse"] = out["BestAvgResponse"].apply(categorize_best_avg_response)
    out["Response_Best_Response"] = out["BestResponse"].apply(categorize_best_response)
    out["Response_BestAvgResponse"] = out["Response_BestAvgResponse"].astype(str)
    out["Response_Best_Response"] = out["Response_Best_Response"].astype(str)
    return out


def assign_durable_column(df: pd.DataFrame, threshold_days: float = DCB_DAYS_THRESHOLD) -> pd.DataFrame:
    """Classify durable clinical benefit from TimeToDouble."""
    out = df.copy()
    out["Durable"] = np.where(out["TimeToDouble"] > threshold_days, "DCB", "non_DCB")
    out["Durable"] = out["Durable"].astype(str)
    return out


def assign_dense_level_groups(
    df: pd.DataFrame,
    *,
    model_col: str = "Model",
    level_col: str = "LEVEL",
    group_col: str = "GROUP",
    rank_pool_from: int = RANK_POOL_FROM,
    pooled_label: str = RANK_POOLED_LABEL,
) -> pd.DataFrame:
    """Assign rank groups from dense LEVEL ordering within each patient."""
    out = df.copy()
    level_rank = (
        out.groupby(model_col, group_keys=False)[level_col]
        .rank(method="dense", ascending=False)
        .astype(int)
    )
    rank_pooled = level_rank.clip(upper=rank_pool_from)
    out[group_col] = [
        pooled_label if r >= rank_pool_from else str(int(r))
        for r in rank_pooled
    ]
    return out


def load_main_source() -> pd.DataFrame:
    """Load single-agent PDX records with valid DDA LEVEL scores."""
    df = pd.read_excel(SOURCE_XLSX)
    df = df[df["LEVEL"] != "#HIÁNYZIK"].copy()
    df["LEVEL"] = pd.to_numeric(df["LEVEL"], errors="coerce")
    df["TimeToDouble"] = pd.to_numeric(df["TimeToDouble"], errors="coerce")
    df = df[df["Treatment type"] == TREAT_TYPE].copy()
    return df


def normalize_model_compound(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "Model" in out.columns:
        out["Model"] = out["Model"].astype(str).str.replace("-", "", regex=False).str.lower()
    if "COMPOUND" in out.columns:
        out["COMPOUND"] = out["COMPOUND"].astype(str).str.replace("-", "", regex=False).str.lower()
    return out


def harmonize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Align column names with the shared PDX analysis schema."""
    out = df.copy()
    if "Treatment target" not in out.columns and "TARGET" in out.columns:
        out = out.rename(columns={"TARGET": "Treatment target"})
    if "COMPOUND" not in out.columns and "Treatment" in out.columns:
        out = out.rename(columns={"Treatment": "COMPOUND"})
    return out


def load_chemo_reference() -> pd.DataFrame:
    """Load standard-of-care chemotherapy reference cohort."""
    chemo_list = ["chemotherapy", "Tubulin"]
    if CHEMO_XLSX.is_file():
        chemo = pd.read_excel(CHEMO_XLSX, sheet_name="simple")
    else:
        chemo = harmonize_column_names(pd.read_csv(CHEMO_CSV))
        chemo = chemo[chemo["Treatment target"].isin(chemo_list)].copy()
        if "molprofil" in chemo.columns:
            chemo = chemo[chemo["molprofil"] != "#HIÁNYZIK"]
    chemo = normalize_model_compound(chemo)
    chemo["CENSOR"] = "1"
    chemo["GROUP"] = "SC"
    chemo = assign_mrecist_columns(chemo)
    chemo = assign_durable_column(chemo)
    return chemo


# ---------------------------------------------------------------------------
# Cohort construction
# ---------------------------------------------------------------------------


def build_on_label_cohort() -> Tuple[pd.DataFrame, List[str], List[str]]:
    """On-label utility cohort: LEVEL tiers with on/off-label subgroups."""
    patient_data = normalize_model_compound(load_main_source())
    patient_data = assign_mrecist_columns(patient_data)
    patient_data = assign_durable_column(patient_data)

    chemo = load_chemo_reference()

    on_col = "On_label" if "On_label" in patient_data.columns else "On_off_label"
    on_yes = "IGAZ" if on_col == "On_label" else "YES_on"
    on_no = "HAMIS" if on_col == "On_label" else "no_off"

    r1 = patient_data[patient_data["LEVEL"] < 0].copy()
    r2 = patient_data[(patient_data["LEVEL"] >= 0) & (patient_data["LEVEL"] < 1000)].copy()
    r3 = patient_data[patient_data["LEVEL"] > 1000].copy()
    r4 = patient_data[
        (patient_data["LEVEL"] > 1000)
        & (patient_data[on_col] == on_yes)
        & (patient_data["APPROVED"] == "IGAZ")
    ].copy()
    r5 = patient_data[
        (patient_data["LEVEL"] > 1000)
        & (patient_data[on_col] == on_no)
        & (patient_data["APPROVED"] == "IGAZ")
    ].copy()
    r6 = patient_data[(patient_data["LEVEL"] > 1000) & (patient_data[on_col] == on_no)].copy()

    labels = {
        id(r1): "Low",
        id(r2): "Int",
        id(r3): "High",
        id(r4): "High On",
        id(r5): "High Off",
        id(r6): "High Off+Exp",
    }
    for frame in (r1, r2, r3, r4, r5, r6):
        frame["GROUP"] = labels[id(frame)]

    all_arm = patient_data.copy()
    all_arm["GROUP"] = "all"

    combined = pd.concat([r1, r2, r3, r4, r5, r6, all_arm, chemo], ignore_index=True)
    summary_groups = ["Low", "Int", "High", "High On", "High Off", "High Off+Exp", "all", "SC"]
    fisher_groups = summary_groups
    return combined, summary_groups, fisher_groups


def build_ranking_cohort() -> Tuple[pd.DataFrame, List[str], List[str]]:
    """Ranking cohort: dense per-patient LEVEL rank with SC chemotherapy reference."""
    df = load_main_source()
    df = assign_mrecist_columns(df)
    df = assign_durable_column(df)

    chemo_list = ["chemotherapy", "Tubulin"]
    chemo_src = harmonize_column_names(pd.read_csv(CHEMO_CSV))
    chemo_src["CENSOR"] = "1"
    target_col = "Treatment target" if "Treatment target" in chemo_src.columns else "TARGET"
    chemo = chemo_src[chemo_src[target_col].isin(chemo_list)].copy()
    chemo["GROUP"] = "SC"
    if "molprofil" in chemo.columns:
        chemo = chemo[chemo["molprofil"] != "#HIÁNYZIK"]
    chemo = assign_mrecist_columns(chemo)
    chemo = assign_durable_column(chemo)

    target_col_pdx = "TARGET" if "TARGET" in df.columns else "Treatment target"
    pdx = normalize_model_compound(df[~df[target_col_pdx].isin(chemo_list)].copy())
    ranked = assign_dense_level_groups(pdx, model_col="Model", level_col="LEVEL", group_col="GROUP")
    all_arm = pdx.copy()
    all_arm["GROUP"] = "all"
    combined = pd.concat([ranked, all_arm, chemo], ignore_index=True)
    rank_labels = ["1", "2", "3", "4", "5", "6", "7-10"]
    summary_groups = rank_labels + ["all", "SC"]
    fisher_groups = rank_labels + ["SC"]
    return combined, summary_groups, fisher_groups


def build_shiva_cohort() -> Tuple[pd.DataFrame, List[str], List[str]]:
    """SHIVA cohort: Low / Intermediate / High tiers by LEVEL."""
    df = load_main_source()
    target_col = "TARGET" if "TARGET" in df.columns else "Treatment target"
    df = df[~df[target_col].str.contains(r"\bchemotherapy\b", case=False, na=False)]
    df = df[~df[target_col].str.contains(r"\bTubulin\b", case=False, na=False)]
    df = assign_mrecist_columns(df)
    df = assign_durable_column(df)

    g_low = df[df["LEVEL"] < 0].copy()
    g_int = df[(df["LEVEL"] >= 0) & (df["LEVEL"] < 1000)].copy()
    g_high = df[df["LEVEL"] > 1000].copy()
    g_low["GROUP"] = "Low"
    g_int["GROUP"] = "Intermediate"
    g_high["GROUP"] = "High"

    all_arm = df.copy()
    all_arm["GROUP"] = "all"
    combined = pd.concat([g_low, g_int, g_high, all_arm], ignore_index=True)
    summary_groups = ["Low", "Intermediate", "High", "all"]
    fisher_groups = summary_groups
    return combined, summary_groups, fisher_groups


GROUPING_BUILDERS = {
    "on_label": build_on_label_cohort,
    "ranking": build_ranking_cohort,
    "shiva": build_shiva_cohort,
}

# ---------------------------------------------------------------------------
# Summary tables
# ---------------------------------------------------------------------------


def compute_dcr_orr_summary(
    data: pd.DataFrame, response_col: str, groups: Sequence[str]
) -> pd.DataFrame:
    rows = []
    for group in groups:
        sub = data[data["GROUP"] == group]
        n = len(sub)
        pd_n = (sub[response_col] == "PD").sum()
        pr_n = (sub[response_col] == "PR").sum()
        cr_n = (sub[response_col] == "CR").sum()
        sd_n = (sub[response_col] == "SD").sum()
        if n == 0:
            orr_pct, dcr_pct = 0.0, 0.0
        else:
            orr_pct = round((pr_n + cr_n) / n * 100, 1)
            dcr_pct = round((cr_n + pr_n + sd_n) / n * 100, 1)
        rows.append(
            {
                "GROUP": group,
                "ALL": n,
                "PD": pd_n,
                "PR": pr_n,
                "CR": cr_n,
                "SD": sd_n,
                "ORR": orr_pct,
                "DCR": dcr_pct,
                "exp_ORR": pr_n + cr_n,
                "control_ORR": sd_n + pd_n,
                "exp_DCR": cr_n + pr_n + sd_n,
                "control_DCR": pd_n,
            }
        )
    return pd.DataFrame(rows)


def compute_db_summary(data: pd.DataFrame, groups: Sequence[str]) -> pd.DataFrame:
    rows = []
    for group in groups:
        sub = data[data["GROUP"] == group]
        n = len(sub)
        dcb_n = (sub["Durable"] == "DCB").sum()
        non_dcb_n = (sub["Durable"] == "non_DCB").sum()
        if n == 0:
            dcb_pct, non_dcb_pct = 0.0, 0.0
        else:
            dcb_pct = round(dcb_n / n * 100, 0)
            non_dcb_pct = round(non_dcb_n / n * 100, 0)
        rows.append(
            {
                "GROUP": group,
                "ALL": n,
                "DB (%)": dcb_pct,
                "non_DB (%)": non_dcb_pct,
                "number_DB": dcb_n,
                "number_non_DB": non_dcb_n,
                "exp_DB": dcb_n,
                "control_DB": non_dcb_n,
            }
        )
    return pd.DataFrame(rows)


def melt_dcr_orr_for_global(
    summary: pd.DataFrame, grouping_method: str, response_classification: str
) -> pd.DataFrame:
    long_rows = []
    for _, row in summary.iterrows():
        for endpoint in ENDPOINTS_ORR_DCR:
            long_rows.append(
                {
                    "grouping_method": grouping_method,
                    "response_classification": response_classification,
                    "response_classification_label": RESPONSE_CLASS_LABEL[response_classification],
                    "GROUP": row["GROUP"],
                    "endpoint": endpoint,
                    "metric_type": endpoint,
                    "percent": row[endpoint],
                    "n_numerator": row[f"exp_{endpoint}"],
                    "n_denominator": row["ALL"],
                    "ALL": row["ALL"],
                    "PD": row["PD"],
                    "PR": row["PR"],
                    "CR": row["CR"],
                    "SD": row["SD"],
                }
            )
    return pd.DataFrame(long_rows)


def melt_db_for_global(summary: pd.DataFrame, grouping_method: str) -> pd.DataFrame:
    rows = []
    for _, row in summary.iterrows():
        rows.append(
            {
                "grouping_method": grouping_method,
                "response_classification": "TimeToDouble",
                "response_classification_label": "DCB_42_days",
                "GROUP": row["GROUP"],
                "endpoint": ENDPOINT_DB,
                "metric_type": ENDPOINT_DB,
                "percent": row["DB (%)"],
                "n_numerator": row["number_DB"],
                "n_denominator": row["ALL"],
                "ALL": row["ALL"],
                "number_non_DB": row["number_non_DB"],
            }
        )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Cochran–Armitage trend test
# ---------------------------------------------------------------------------


def cochran_armitage_z_p(successes, failures) -> Tuple[float, float]:
    successes = np.asarray(successes, dtype=float)
    failures = np.asarray(failures, dtype=float)
    totals = successes + failures
    if totals.sum() == 0:
        return 0.0, np.nan
    scores = np.arange(len(successes))
    p_overall = successes.sum() / totals.sum()
    expected = totals * p_overall
    mean_score = np.average(scores, weights=totals)
    numerator = np.sum(scores * (successes - expected))
    denominator = np.sqrt(
        p_overall * (1 - p_overall) * np.sum(totals * (scores - mean_score) ** 2)
    )
    if denominator <= 0:
        return 0.0, np.nan
    z_val = numerator / denominator
    p_val = 2 * (1 - stats.norm.cdf(abs(z_val)))
    return float(z_val), float(p_val)


def compute_cochran_results(
    summary: pd.DataFrame,
    group_order: Sequence[str],
    response_classification: str,
    grouping_method: str,
) -> pd.DataFrame:
    trend_groups = [g for g in group_order if g not in COCHRAN_EXCLUDE_GROUPS]
    cdf = (
        summary[summary["GROUP"].isin(trend_groups)]
        .set_index("GROUP")
        .reindex(trend_groups)
        .reset_index()
    )
    cdf = cdf.dropna(subset=["exp_DCR", "control_DCR"])
    rows = []
    if len(cdf) < 2:
        return pd.DataFrame(rows)

    for endpoint, exp_col, ctrl_col in (
        ("DCR", "exp_DCR", "control_DCR"),
        ("ORR", "exp_ORR", "control_ORR"),
    ):
        z_val, p_val = cochran_armitage_z_p(cdf[exp_col].values, cdf[ctrl_col].values)
        rows.append(
            {
                "grouping_method": grouping_method,
                "response_classification": response_classification,
                "response_classification_label": RESPONSE_CLASS_LABEL.get(
                    response_classification, response_classification
                ),
                "endpoint": endpoint,
                "groups_included": ", ".join(trend_groups),
                "n_groups": len(cdf),
                "Z": z_val,
                "p_value": p_val,
            }
        )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Fisher exact tests and multiplicity correction
# ---------------------------------------------------------------------------


def determine_significance_label(p_value: float) -> str:
    if pd.isna(p_value):
        return ""
    return "Significant" if float(p_value) < 0.05 else "Non-significant"


def compute_fisher_long(
    summary: pd.DataFrame,
    fisher_groups: Sequence[str],
    response_classification: str,
    endpoints: Sequence[str],
    metric_prefix: str = "",
    pairwise_pairs: Optional[Sequence[Tuple[str, str]]] = None,
) -> pd.DataFrame:
    rows = []
    present = set(summary["GROUP"])

    if pairwise_pairs is not None:
        pairs = pairwise_pairs
    else:
        present_list = [g for g in fisher_groups if g in present]
        pairs = list(combinations(present_list, 2))

    for endpoint in endpoints:
        if metric_prefix:
            exp_col, ctrl_col = f"exp_{metric_prefix}", f"control_{metric_prefix}"
        else:
            exp_col, ctrl_col = f"exp_{endpoint}", f"control_{endpoint}"

        for g1, g2 in pairs:
            if g1 not in present or g2 not in present:
                continue
            s1 = summary[summary["GROUP"] == g1]
            s2 = summary[summary["GROUP"] == g2]
            if len(s1) != 1 or len(s2) != 1:
                continue
            table = [
                [int(s1.iloc[0][exp_col]), int(s1.iloc[0][ctrl_col])],
                [int(s2.iloc[0][exp_col]), int(s2.iloc[0][ctrl_col])],
            ]
            odds_ratio, p_value = fisher_exact(table)
            rows.append(
                {
                    "response_classification": response_classification,
                    "response": endpoint,
                    "group1": g1,
                    "group2": g2,
                    "Comparison": f"{g1} vs. {g2}",
                    "odds_ratio": odds_ratio,
                    "p_value": p_value,
                }
            )
    return pd.DataFrame(rows)


def apply_fdr(
    fisher_df: pd.DataFrame,
    group_cols: Sequence[str] = ("response_classification", "response"),
    p_col: str = "p_value",
) -> pd.DataFrame:
    out = fisher_df.copy()
    out[p_col] = pd.to_numeric(out[p_col], errors="coerce")
    out["p_fdr"] = np.nan
    for _, block in out.groupby(list(group_cols)):
        valid = block[p_col].dropna().index
        if len(valid) == 0:
            continue
        pvals = out.loc[valid, p_col].astype(float)
        _, p_fdr, _, _ = multipletests(pvals.values, method="fdr_bh")
        out.loc[valid, "p_fdr"] = p_fdr
    out["Sig_raw"] = out[p_col].apply(determine_significance_label)
    out["Sig_fdr"] = out["p_fdr"].apply(determine_significance_label)
    out["p_raw"] = out[p_col]
    return out


def fisher_summary_sheet(fisher_corrected: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "response_classification",
        "response_classification_label",
        "response",
        "group1",
        "group2",
        "Comparison",
        "odds_ratio",
        "p_raw",
        "p_fdr",
        "Sig_raw",
        "Sig_fdr",
    ]
    out = fisher_corrected.copy()
    if "response_classification_label" not in out.columns:
        out["response_classification_label"] = out["response_classification"].map(
            RESPONSE_CLASS_LABEL
        ).fillna(out["response_classification"])
    existing = [c for c in cols if c in out.columns]
    return out[existing]


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------


def _is_pvalue_column(name) -> bool:
    if name is None:
        return False
    n = str(name).lower()
    return n in {"p_raw", "p_fdr", "p_value"} or n.startswith("p value")


def apply_excel_formats(path: Path, sheet_names: Optional[Iterable[str]] = None) -> None:
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    wb = load_workbook(path)
    names = sheet_names or wb.sheetnames
    for ws_name in names:
        if ws_name not in wb.sheetnames:
            continue
        ws = wb[ws_name]
        header = [c.value for c in ws[1]]
        for row_idx in range(2, ws.max_row + 1):
            for col_idx, col_name in enumerate(header, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_name in {"Sig_raw", "Sig_fdr"}:
                    if cell.value == "Significant":
                        cell.fill = green
                    elif cell.value == "Non-significant":
                        cell.fill = yellow
                elif _is_pvalue_column(col_name):
                    if cell.value in (None, ""):
                        continue
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = SCI_NUMBER_FORMAT
                        cell.fill = green if cell.value < 0.05 else yellow
                    except (TypeError, ValueError):
                        pass
                elif str(col_name).lower() == "odds_ratio":
                    if cell.value not in (None, ""):
                        try:
                            cell.value = float(cell.value)
                            cell.number_format = ODDS_RATIO_NUMBER_FORMAT
                        except (TypeError, ValueError):
                            pass
    wb.save(path)


def save_excel(path: Path, sheets: dict, apply_format: bool = True) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
    if apply_format:
        apply_excel_formats(path)


# ---------------------------------------------------------------------------
# Pipeline execution
#
# Flow per stratification method:
#   1. Build cohort and assign comparison groups.
#   2. Summarise ORR, DCR, and DB by group.
#   3. Pairwise Fisher exact tests with BH-FDR correction.
#   4. Cochran–Armitage trend test over ordered tiers.
#   5. Export per-method and combined Excel workbooks.
# ---------------------------------------------------------------------------


def _sha256_of_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _file_meta(path: Path) -> dict:
    meta = {"path": str(path), "exists": path.is_file()}
    if path.is_file():
        meta["sha256"] = _sha256_of_file(path)
        meta["size_bytes"] = path.stat().st_size
    return meta


def write_run_metadata(methods_run: Sequence[str]) -> None:
    """Record input checksums and run configuration for reproducibility."""
    metadata = {
        "generated_at": datetime.now().isoformat(),
        "run_date": RUN_DATE,
        "output_root": str(OUTPUT_ROOT),
        "analyses_run": list(methods_run),
        "source": _file_meta(SOURCE_XLSX),
        "chemo_csv": _file_meta(CHEMO_CSV),
        "chemo_xlsx": _file_meta(CHEMO_XLSX),
        "dcb_days_threshold": DCB_DAYS_THRESHOLD,
        "rank_pool_from": RANK_POOL_FROM,
    }
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    (OUTPUT_ROOT / "run_metadata.json").write_text(
        json.dumps(metadata, indent=2), encoding="utf-8"
    )


def _fdr_count_summary(fisher_df: pd.DataFrame) -> pd.DataFrame:
    if fisher_df.empty:
        return pd.DataFrame()
    group_cols = ["response_classification", "response"]
    if "response_classification_label" in fisher_df.columns:
        group_cols.append("response_classification_label")
    return (
        fisher_df.groupby(group_cols, dropna=False)
        .agg(
            n_tests=("p_raw", "count"),
            n_raw_significant=("Sig_raw", lambda s: (s == "Significant").sum()),
            n_fdr_significant=("Sig_fdr", lambda s: (s == "Significant").sum()),
        )
        .reset_index()
    )


def run_grouping_method(method_name: str) -> None:
    """Execute the full DCR / ORR / DB workflow for one stratification scheme."""
    builder = GROUPING_BUILDERS[method_name]
    data, summary_groups, fisher_groups = builder()
    method_dir = OUTPUT_ROOT / method_name
    dcr_orr_dir = method_dir / "DCR_ORR"
    durable_dir = method_dir / "durable"
    dcr_orr_dir.mkdir(parents=True, exist_ok=True)
    durable_dir.mkdir(parents=True, exist_ok=True)

    global_metric_rows: List[pd.DataFrame] = []
    all_dcr_orr_summaries: List[pd.DataFrame] = []
    all_dcr_orr_fisher: List[pd.DataFrame] = []
    db_summary = compute_db_summary(data, summary_groups)
    db_summary["grouping_method"] = method_name

    fisher_pairs = RANKING_FISHER_PAIRS if method_name == "ranking" else None

    for resp_col in RESPONSE_CLASSIFICATIONS:
        summary = compute_dcr_orr_summary(data, resp_col, summary_groups)
        summary["response_classification"] = resp_col
        summary["response_classification_label"] = RESPONSE_CLASS_LABEL[resp_col]
        summary["grouping_method"] = method_name
        all_dcr_orr_summaries.append(summary)

        global_metric_rows.append(melt_dcr_orr_for_global(summary, method_name, resp_col))

        fisher_raw = compute_fisher_long(
            summary, fisher_groups, resp_col, ENDPOINTS_ORR_DCR,
            pairwise_pairs=fisher_pairs,
        )
        fisher_corr = apply_fdr(fisher_raw)
        fisher_corr["response_classification_label"] = RESPONSE_CLASS_LABEL[resp_col]
        fisher_corr["grouping_method"] = method_name
        all_dcr_orr_fisher.append(fisher_corr)

        tag = RESPONSE_CLASS_LABEL[resp_col]
        save_excel(
            dcr_orr_dir / f"{RUN_DATE}_DCR_ORR_summary_{tag}.xlsx",
            {"Summary": summary, "Fisher_FDR": fisher_summary_sheet(fisher_corr)},
        )

    global_metric_rows.append(melt_db_for_global(db_summary, method_name))

    fisher_db_raw = compute_fisher_long(
        db_summary, fisher_groups, "TimeToDouble", [ENDPOINT_DB],
        metric_prefix="DB", pairwise_pairs=fisher_pairs,
    )
    fisher_db_corr = apply_fdr(
        fisher_db_raw,
        group_cols=("response_classification", "response"),
    )
    fisher_db_corr["grouping_method"] = method_name
    fisher_db_corr["response_classification_label"] = "DCB_42_days"

    save_excel(
        durable_dir / f"{RUN_DATE}_DB_summary.xlsx",
        {"Summary": db_summary, "Fisher_FDR": fisher_summary_sheet(fisher_db_corr)},
    )

    combined_dcr_orr_summary = pd.concat(all_dcr_orr_summaries, ignore_index=True)
    combined_dcr_orr_fisher = pd.concat(all_dcr_orr_fisher, ignore_index=True)
    global_metrics = pd.concat(global_metric_rows, ignore_index=True)

    cochran_order = COCHRAN_GROUP_ORDER[method_name]
    cochran_blocks = [
        compute_cochran_results(s, cochran_order, rc, method_name)
        for s, rc in zip(all_dcr_orr_summaries, RESPONSE_CLASSIFICATIONS)
    ]
    cochran_combined = pd.concat(cochran_blocks, ignore_index=True)

    save_excel(
        dcr_orr_dir / f"{RUN_DATE}_DCR_ORR_combined_both_classifications.xlsx",
        {
            "Summary_combined": combined_dcr_orr_summary,
            "Fisher_FDR_combined": fisher_summary_sheet(combined_dcr_orr_fisher),
        },
    )

    global_path = method_dir / f"DCR_ORR_DB_{method_name}_global_summary.xlsx"
    save_excel(
        global_path,
        {
            "Metrics_long": global_metrics,
            "DCR_ORR_by_group": combined_dcr_orr_summary,
            "DB_by_group": db_summary,
            "Fisher_DCR_ORR": fisher_summary_sheet(combined_dcr_orr_fisher),
            "Fisher_DB": fisher_summary_sheet(fisher_db_corr),
            "FDR_summary_DCR_ORR": _fdr_count_summary(combined_dcr_orr_fisher),
            "FDR_summary_DB": _fdr_count_summary(fisher_db_corr),
            "Cochran_Armitage": cochran_combined,
        },
    )
    LOGGER.info("[%s] Wrote %s", method_name, global_path)


def write_run_readme() -> None:
    readme = OUTPUT_ROOT / "README_pipeline.txt"
    text = f"""Global DCR / ORR / DB pipeline
Run date: {RUN_DATE}
Source: {SOURCE_XLSX}

Subfolders:
  on_label/  — LEVEL tiers with on/off-label splits and SC chemotherapy reference
  ranking/   — per-patient dense LEVEL rank (1–7-10; ties share group) and SC reference
  shiva/     — Low / Intermediate / High by LEVEL (chemotherapy excluded from main cohort)

Each method folder contains:
  DCR_ORR_DB_{{method}}_global_summary.xlsx — DCR, ORR, and DB for both mRECIST classifications
  DCR_ORR/ — per-classification and combined Fisher tests with BH-FDR correction
  durable/ — durable clinical benefit (DB) at TimeToDouble > {DCB_DAYS_THRESHOLD} days

Ranking pairwise comparisons: rank 1 vs ranks 2, 3, 4, 5, 6, 7-10, and SC.

Response classification columns:
  Response_Best_Response   — mRECIST from BestResponse
  Response_BestAvgResponse — mRECIST from BestAvgResponse
  TimeToDouble             — durable (DB) endpoint only
"""
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    readme.write_text(text, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Global DCR / ORR / DB analysis for on-label, ranking, and SHIVA cohorts."
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=DEFAULT_SOURCE,
        help="Primary PDX input workbook (default: data/pdx_curve_metrics_single_treatments_dda_scores.xlsx).",
    )
    parser.add_argument(
        "--chemo-csv",
        type=Path,
        default=DEFAULT_CHEMO_CSV,
        help="Chemotherapy reference cohort CSV for ranking analysis.",
    )
    parser.add_argument(
        "--chemo-xlsx",
        type=Path,
        default=DEFAULT_CHEMO_XLSX,
        help="Chemotherapy reference workbook for on-label analysis.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_ROOT,
        help="Directory for analysis outputs.",
    )
    parser.add_argument(
        "--analyses",
        nargs="+",
        choices=["all", *ANALYSIS_METHODS],
        default=["all"],
        help="Stratification methods to run (default: all).",
    )
    return parser.parse_args()


def main() -> None:
    global SOURCE_XLSX, CHEMO_CSV, CHEMO_XLSX, OUTPUT_ROOT

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
    )

    args = parse_args()
    SOURCE_XLSX = args.source.resolve()
    CHEMO_CSV = args.chemo_csv.resolve()
    CHEMO_XLSX = args.chemo_xlsx.resolve()
    OUTPUT_ROOT = args.output_dir.resolve()

    LOGGER.info("Source: %s", SOURCE_XLSX)
    LOGGER.info("Output: %s", OUTPUT_ROOT)

    if not SOURCE_XLSX.is_file():
        raise FileNotFoundError(f"Source file not found: {SOURCE_XLSX}")
    if not CHEMO_CSV.is_file() and not CHEMO_XLSX.is_file():
        raise FileNotFoundError(
            "Chemotherapy reference not found. Expected "
            f"{CHEMO_CSV} or {CHEMO_XLSX}"
        )

    write_run_readme()

    selected = (
        list(ANALYSIS_METHODS)
        if "all" in args.analyses
        else [m for m in args.analyses if m in ANALYSIS_METHODS]
    )
    for method in selected:
        run_grouping_method(method)

    write_run_metadata(selected)
    LOGGER.info("Done. Outputs: %s", OUTPUT_ROOT)


if __name__ == "__main__":
    main()
