"""Kaplan-Meier analysis pipeline for PDX treatment cohorts.

Runs three analysis branches equivalent to the legacy scripts:
  1. On-label / off-label DDA tier comparison (Low, Int, High, On, Off,
     Off+Ex vs. standard-of-care chemotherapy).
  2. Per-patient compound ranking (dense rank 1–6, pooled 7-10 vs. SC).
  3. SHIVA-tier comparison (Low, Int, High, All, Non-assigned, SC).

All branches share a single primary input workbook located in data/.

Usage
-----
python KM_PDX.py [--source PATH] [--output-dir DIR]
                 [--analyses {on_label,ranking,shiva,all} ...]

Random seed: SEED = 42 (set at module level via np.random.seed).
Outputs are written to <output-dir>/KM_PDX/ with per-analysis subfolders.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from lifelines import CoxPHFitter, KaplanMeierFitter
from lifelines.plotting import add_at_risk_counts
from lifelines.statistics import logrank_test, pairwise_logrank_test
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from statsmodels.stats.multitest import multipletests


# ---------------------------------------------------------------------------
# Reproducibility
# ---------------------------------------------------------------------------
SEED = 42
np.random.seed(SEED)

LOGGER = logging.getLogger("KM_PDX")
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")

# ---------------------------------------------------------------------------
# Repository-relative paths (no hard-coded absolute paths)
# ---------------------------------------------------------------------------
_SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = _SCRIPT_DIR.parent
DATA_DIR = REPO_ROOT / "data"

DEFAULT_DATA_COLUMN = "TimeToDouble"
KM_XLIM = (0, 200)
KM_XTICKS = np.arange(0, 225, 25)

DEFAULT_SOURCE_XLSX = (
    DATA_DIR
    / "pdx_curve_metrics_single_treatments_dda_scores.xlsx"
)

KM_PDX_OUTPUT_DIRNAME = "KM_PDX"


# =============================================================================
# Embedded helper: pairwise log-rank Excel export with FDR highlighting
# (originally in codes/km_pairwise_excel_export.py)
# =============================================================================

_FILL_CHANGED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_UNCHANGED = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_SIG = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_NS = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
_FILL_CHANGED_TRUE = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
_FILL_CHANGED_FALSE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

_SIG_COLS = ("raw_significance", "fdr_significance")
_PAIRWISE_GROUP_RENAME = {"level_0": "Group_1", "level_1": "Group_2"}


def _rename_pairwise_group_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename = {k: v for k, v in _PAIRWISE_GROUP_RENAME.items() if k in df.columns}
    return df.rename(columns=rename) if rename else df


def _sig_is_significant(label: Any) -> bool:
    """Return True if the significance label represents any star level."""
    return str(label).strip().lower() not in ("", "ns")


def _format_p_value(p_value: Any) -> str:
    return f"{float(p_value):.2e}"


def _reorder_p_adj_columns(df: pd.DataFrame) -> pd.DataFrame:
    _p_adj_cols = ("p_fdr_bh", "raw_significance", "fdr_significance")
    out = df.copy()
    _cols = list(out.columns)
    for _c in _p_adj_cols:
        if _c in _cols:
            _cols.remove(_c)
    if "p" in _cols:
        _insert_at = _cols.index("p") + 1
        for _c in _p_adj_cols:
            if _c in out.columns:
                _cols.insert(_insert_at, _c)
                _insert_at += 1
        out = out[_cols]
    return out


def _add_fdr_significance_changed(df: pd.DataFrame) -> pd.DataFrame:
    """Annotate rows where FDR correction changes significance at alpha=0.05."""
    out = df.copy()
    if "raw_significance" not in out.columns or "fdr_significance" not in out.columns:
        return out
    out["fdr_significance_changed"] = [
        "changed" if _sig_is_significant(r) != _sig_is_significant(f) else "unchanged"
        for r, f in zip(out["raw_significance"], out["fdr_significance"])
    ]
    return out


def _reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = list(df.columns)
    if "fdr_significance_changed" in cols:
        cols.remove("fdr_significance_changed")
    if "fdr_significance" in cols:
        cols.insert(cols.index("fdr_significance") + 1, "fdr_significance_changed")
    return df[cols]


def export_pairwise_logrank_excel(df: pd.DataFrame, path: str) -> None:
    """Write pairwise log-rank table with colour-coded FDR significance changes."""
    out = _reorder_columns(_add_fdr_significance_changed(df))
    out.to_excel(path, index=False, engine="openpyxl")
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if "fdr_significance_changed" not in headers:
        wb.save(path)
        return
    col_idx = headers.index("fdr_significance_changed") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value == "changed":
            cell.fill = _FILL_CHANGED
        elif cell.value == "unchanged":
            cell.fill = _FILL_UNCHANGED
    wb.save(path)


# =============================================================================
# Embedded helper: dense per-patient LEVEL ranking
# (originally in codes/ranking_dense_level.py)
# =============================================================================

_RANK_POOL_FROM = 7
_POOLED_LABEL = "7-10"


def assign_dense_level_groups(
    df: pd.DataFrame,
    *,
    model_col: str = "Model",
    level_col: str = "LEVEL",
    group_col: str = "GROUP",
    rank_pool_from: int = _RANK_POOL_FROM,
    pooled_label: str = _POOLED_LABEL,
) -> pd.DataFrame:
    """Assign GROUP from dense rank of level_col within each model_col group.

    Higher LEVEL → lower rank number (1 = best DDA response). Tied values
    share rank. Ranks >= rank_pool_from are pooled into pooled_label ('7-10').
    """
    out = df.copy()
    level_rank = (
        out.groupby(model_col, group_keys=False)[level_col]
        .rank(method="dense", ascending=False)
        .astype(int)
    )
    rank_pooled = level_rank.clip(upper=rank_pool_from)
    out[group_col] = [
        pooled_label if r >= rank_pool_from else str(int(r))
        for r in rank_pooled
    ]
    return out


# =============================================================================
# Configuration containers
# =============================================================================


@dataclass(frozen=True)
class Paths:
    """Filesystem paths resolved relative to the repository data/ directory."""

    project_root: Path
    script_dir: Path
    source_primary_xlsx: Path
    chemo_csv: Path
    non_assigned_xlsx: Path


@dataclass(frozen=True)
class KMAnalysisSpec:
    """Metadata for one Kaplan-Meier analysis branch."""

    run_key: str
    source_script_stem: str
    legacy_name: str
    date: str
    data_column: str = DEFAULT_DATA_COLUMN
    primary_tumor: str = ""
    pair_comparisons: Tuple[Tuple[str, str], ...] = ()
    extra: Dict[str, Any] = field(default_factory=dict)


@dataclass
class KMPlotStyle:
    """Shared typography parameters (scaled from legacy 6.4 pt base)."""

    font: float = 6.4 * 1.6
    tick_size: float = 6.4 * 1.6
    label_size: float = 6.4 * 1.6
    at_risk: float = 5.5 * 1.6
    font_legend: float = 6.4 * 1.6
    cishow: bool = False

    def apply_rcparams(self) -> None:
        plt.rcParams.update(
            {
                "axes.titlesize": self.font,
                "axes.labelsize": self.font,
                "xtick.labelsize": self.at_risk,
                "ytick.labelsize": self.font,
                "legend.fontsize": self.font_legend,
                "font.size": self.font,
            }
        )


# =============================================================================
# CLI and path resolution
# =============================================================================


def parse_args() -> argparse.Namespace:
    """Parse command-line options."""
    parser = argparse.ArgumentParser(
        description="Kaplan-Meier analysis for PDX treatment cohorts (on-label, ranking, SHIVA)."
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=None,
        help=(
            f"Primary input workbook override "
            f"(default: data/{DEFAULT_SOURCE_XLSX.name})."
        ),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=_SCRIPT_DIR,
        help=f"Parent output directory; results in <output-dir>/{KM_PDX_OUTPUT_DIRNAME}/.",
    )
    parser.add_argument(
        "--analyses",
        nargs="+",
        choices=("on_label", "ranking", "shiva", "all"),
        default=["all"],
        help="Analyses to execute (default: all).",
    )
    return parser.parse_args()


def filesystem_safe_name(text: str) -> str:
    allowed = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_")
    safe = "".join(ch if ch in allowed else "_" for ch in text.strip().replace(" ", "_"))
    while "__" in safe:
        safe = safe.replace("__", "_")
    return safe.strip("_")


def sha256_of_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def file_meta(path: Path) -> Dict[str, str]:
    meta: Dict[str, str] = {"path": str(path), "exists": str(path.exists())}
    if path.exists():
        meta["sha256"] = sha256_of_file(path)
        meta["size_bytes"] = str(path.stat().st_size)
    return meta


def legacy_output_folder_name(spec: KMAnalysisSpec) -> str:
    return f"{spec.date}{spec.legacy_name}_{spec.data_column}_"


def analysis_subfolder_name(spec: KMAnalysisSpec) -> str:
    fixed_names = {
        "ranking": "KM_ranking",
        "shiva": "KM_SHIVA",
        "on_label": "KM_utility",
    }
    return fixed_names.get(
        spec.run_key,
        f"{filesystem_safe_name(spec.source_script_stem)}__{filesystem_safe_name(spec.legacy_name)}",
    )


def resolve_analysis_output_dir(output_root: Path, spec: KMAnalysisSpec) -> Path:
    nested = output_root / analysis_subfolder_name(spec) / legacy_output_folder_name(spec)
    nested.mkdir(parents=True, exist_ok=True)
    return nested


def write_output_pointer(script_dir: Path, output_dir: Path) -> None:
    pointer = script_dir / "KM_PDX_output_folder.txt"
    pointer.write_text(str(output_dir.resolve()), encoding="utf-8")
    LOGGER.info("Output pointer written to %s", pointer)


# =============================================================================
# Shared data preparation
# =============================================================================


def normalize_id_columns(
    df: pd.DataFrame, model_col: str = "Model", compound_col: str = "COMPOUND"
) -> pd.DataFrame:
    """Lowercase and strip hyphens from model/compound identifiers."""
    out = df.copy()
    out[model_col] = out[model_col].astype(str).str.replace("-", "", regex=False).str.lower()
    out[compound_col] = out[compound_col].astype(str).str.replace("-", "", regex=False).str.lower()
    return out


def harmonize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Map legacy TARGET/Treatment column names to expected downstream names."""
    out = df.copy()
    if "Treatment target" not in out.columns and "TARGET" in out.columns:
        out = out.rename(columns={"TARGET": "Treatment target"})
    if "Treatment" in out.columns and "COMPOUND" not in out.columns:
        out = out.rename(columns={"Treatment": "COMPOUND"})
    return out


def clean_base_patient_table(df: pd.DataFrame) -> pd.DataFrame:
    """Apply shared row filtering and numeric casting."""
    out = df.copy()
    out = out[out["LEVEL"] != "#HIÁNYZIK"].copy()
    out["LEVEL"] = out["LEVEL"].astype(float)
    out["TimeToDouble"] = pd.to_numeric(out["TimeToDouble"], errors="coerce").astype(float)
    if "Day_Last" in out.columns:
        out["Day_Last"] = pd.to_numeric(out["Day_Last"], errors="coerce")
    return out


def normalize_survival_columns(
    df: pd.DataFrame, time_col: str, event_col: str = "CENSOR"
) -> pd.DataFrame:
    """Ensure lifelines receives numeric durations and boolean event indicators."""
    out = df.copy()
    out.loc[:, time_col] = pd.to_numeric(out[time_col], errors="coerce")
    ev = out[event_col].replace({True: 1, False: 0, "1": 1, "0": 0, "True": 1, "False": 0})
    events = pd.to_numeric(ev, errors="coerce").fillna(0).astype(int).to_numpy(dtype=np.bool_, copy=False)
    out[event_col] = pd.Series(events, index=out.index, dtype=bool)
    return out


def validate_on_label_source_columns(patient_df: pd.DataFrame) -> None:
    missing = {"APPROVED", "On_label"}.difference(patient_df.columns)
    if missing:
        raise KeyError(
            f"Source table is missing on-label columns: {sorted(missing)}. "
            "Expected APPROVED and On_label in the recategorised source workbook."
        )


def resolve_on_label_columns(patient_df: pd.DataFrame) -> Tuple[str, str, str]:
    """Detect on-label column naming convention used in the workbook."""
    on_col = "On_label" if "On_label" in patient_df.columns else "On_off_label"
    on_yes = "IGAZ" if on_col == "On_label" else "YES_on"
    on_no = "HAMIS" if on_col == "On_label" else "no_off"
    return on_col, on_yes, on_no


def load_chemo_from_csv(paths: Paths) -> pd.DataFrame:
    """Load standard-of-care chemotherapy reference cohort from CSV."""
    chemo_list = ["chemotherapy", "Tubulin"]
    chemo = pd.read_csv(paths.chemo_csv)
    chemo = harmonize_column_names(chemo)
    chemo["CENSOR"] = 1
    chemo = chemo[chemo["Treatment target"].isin(chemo_list)].reset_index(drop=True)
    chemo = chemo[chemo["molprofil"] != "#HIÁNYZIK"]
    chemo["GROUP"] = "SC"
    if "COMPOUND" in chemo.columns:
        chemo["COMPOUND"] = chemo["COMPOUND"].astype(str).str.replace("-", "", regex=False).str.lower()
    chemo["Model"] = chemo["Model"].astype(str).str.replace("-", "", regex=False).str.lower()
    return chemo


def load_chemo_from_xlsx(paths: Paths) -> pd.DataFrame:
    """Load on-label chemo cohort from CSV."""
    return load_chemo_from_csv(paths)


def load_non_assigned_table(paths: Paths) -> pd.DataFrame:
    """Load Non_assigned rows; return empty frame with expected columns if file absent."""
    if not paths.non_assigned_xlsx.exists():
        LOGGER.warning(
            "Non_assigned workbook not found at %s; SHIVA subset plot will omit Non_assigned.",
            paths.non_assigned_xlsx,
        )
        return pd.DataFrame(columns=["Model", "COMPOUND", "TimeToDouble", "CENSOR", "GROUP", "LEVEL"])
    non_assigned = pd.read_excel(paths.non_assigned_xlsx)
    non_assigned["CENSOR"] = True
    non_assigned["GROUP"] = "Non_assigned"
    return non_assigned


def validate_configured_sources(paths: Paths) -> Dict[str, Any]:
    primary_meta = file_meta(paths.source_primary_xlsx)
    return {
        "shared_primary_source": str(paths.source_primary_xlsx),
        "shared_primary_source_meta": primary_meta,
        "all_analyses_use_same_source": True,
    }


# =============================================================================
# Statistics and plotting utilities
# =============================================================================


def significance_stars(p_value: float) -> str:
    if not np.isfinite(p_value):
        return "ns"
    if p_value < 0.0005:
        return "***"
    if p_value < 0.005:
        return "**"
    if p_value < 0.05:
        return "*"
    return "ns"


def space_plot_operators(text: str) -> str:
    """Add readable spacing around comparison operators in on-plot annotations."""
    spaced = text
    spaced = re.sub(r">=", " >= ", spaced)
    spaced = re.sub(r"<=", " <= ", spaced)
    spaced = re.sub(r"(?<![<>=!])=(?!=)", " = ", spaced)
    spaced = re.sub(r"(?<![=])>(?!=)", " > ", spaced)
    spaced = re.sub(r"(?<![=!<])<(?!=)", " < ", spaced)
    return re.sub(r"  +", " ", spaced)


def safe_filename(value: str) -> str:
    if not str(value).strip():
        return ""
    text = str(value).strip()
    for ch in '<>:"/\\|?*':
        text = text.replace(ch, "_")
    return text


def apply_km_axis_style(ax: plt.Axes, style: KMPlotStyle, title: str = "") -> None:
    ax.set_xlabel("Time (days)", fontsize=style.label_size, labelpad=4)
    ax.set_ylabel("Survival Probability (%)", fontsize=style.label_size, labelpad=5)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_xlim(*KM_XLIM)
    ax.set_xticks(KM_XTICKS)
    if title:
        ax.set_title(title)


def collect_km_group_statistics(
    kmf: KaplanMeierFitter,
    fit_label: str,
    group: str,
    group_data: pd.DataFrame,
    level_col: str = "LEVEL",
) -> Dict[str, Any]:
    """Extract median survival and 95% CI bounds from a fitted KM curve."""
    lower_bound = kmf.confidence_interval_[f"{fit_label}_lower_0.95"]
    upper_bound = kmf.confidence_interval_[f"{fit_label}_upper_0.95"]
    lower_median_time = np.interp(0.5, 1 - lower_bound.values.flatten(), lower_bound.index)
    upper_median_time = np.interp(0.5, 1 - upper_bound.values.flatten(), upper_bound.index)
    return {
        "Group": group,
        "Median Survival Time": kmf.median_survival_time_,
        "n_number": len(group_data),
        "CI_lower": lower_median_time,
        "CI_upper": upper_median_time,
        "avg_LEVEL": group_data[level_col].mean(),
    }


def fit_multigroup_km_plot(
    patient_data: pd.DataFrame,
    data_column: str,
    colors: Dict[str, str],
    style: KMPlotStyle,
    *,
    legend_labels: Optional[Sequence[str]] = None,
    legend_title: Optional[str] = None,
    line_styles: Optional[Dict[str, str]] = None,
    line_widths: Optional[Dict[str, float]] = None,
    group_order: Optional[Sequence[str]] = None,
    label_suffix_n: bool = False,
    at_risk_ypos: float = -0.35,
) -> Tuple[plt.Figure, plt.Axes, List[KaplanMeierFitter], pd.DataFrame]:
    """Fit and plot Kaplan-Meier curves for all requested groups."""
    style.apply_rcparams()
    fig, ax = plt.subplots(figsize=(5, 5.5))
    kmf_list: List[KaplanMeierFitter] = []
    rows: List[Dict[str, Any]] = []

    groups = list(group_order) if group_order else list(patient_data["GROUP"].unique())
    default_linewidth = 1.0
    line_styles = line_styles or {}
    line_widths = line_widths or {}

    for group in groups:
        group_data = patient_data[patient_data["GROUP"] == group].copy()
        if group_data.empty:
            LOGGER.warning("Skipping empty group: %s", group)
            continue

        group_data.loc[:, data_column] = pd.to_numeric(group_data[data_column], errors="coerce")
        group_data = group_data.dropna(subset=[data_column])
        if group_data.empty:
            continue

        label = f"{group} (n={len(group_data)})" if label_suffix_n else group
        kmf = KaplanMeierFitter()
        kmf.fit(group_data[data_column], event_observed=group_data["CENSOR"], label=label)
        kmf.plot(
            ax=ax,
            color=colors.get(group),
            ci_show=style.cishow,
            linestyle=line_styles.get(group, "-"),
            linewidth=line_widths.get(group, default_linewidth),
        )
        kmf_list.append(kmf)
        stat = collect_km_group_statistics(kmf, label, group, group_data)
        stat["Group"] = group
        rows.append(stat)

    apply_km_axis_style(ax, style)
    if legend_labels is not None:
        ax.legend(labels=list(legend_labels), frameon=False, fontsize=style.font_legend, title=legend_title)
    else:
        ax.legend(frameon=False, fontsize=style.font_legend, title=legend_title)

    if kmf_list:
        add_at_risk_counts(
            *kmf_list,
            ypos=at_risk_ypos,
            ax=ax,
            rows_to_show=["At risk"],
            at_risk_count_from_start_of_period=True,
            xticks=KM_XTICKS,
        )

    plt.subplots_adjust(bottom=0.25, top=0.9, hspace=0.5)
    plt.tight_layout()
    summary_df = pd.DataFrame(rows)
    return fig, ax, kmf_list, summary_df


def append_fdr_to_pairwise_hr(pairwise_hr_df: pd.DataFrame) -> pd.DataFrame:
    """Add BH-FDR adjusted p-values and star labels to a pairwise HR export table."""
    if pairwise_hr_df.empty or not pairwise_hr_df["logrank_p"].notna().any():
        return pairwise_hr_df

    out = pairwise_hr_df.copy()
    raw = out["logrank_p"].astype(float).values
    _, fdr_lr, _, _ = multipletests(raw, method="fdr_bh")
    out["logrank_p_fdr_bh"] = fdr_lr
    out["raw_significance"] = out["logrank_p"].apply(significance_stars)
    out["fdr_significance"] = out["logrank_p_fdr_bh"].apply(significance_stars)

    adj_cols = ("logrank_p_fdr_bh", "raw_significance", "fdr_significance")
    cols = [c for c in out.columns if c not in adj_cols]
    if "logrank_p" in cols:
        insert_at = cols.index("logrank_p") + 1
        for c in adj_cols:
            if c in out.columns:
                cols.insert(insert_at, c)
                insert_at += 1
        out = out[cols]
    return out


def format_pairwise_logrank_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Apply BH-FDR, star labels, and p-value formatting for log-rank export."""
    out = df.reset_index().rename(columns={"index": "GROUP"}) if "GROUP" not in df.columns else df.copy()
    out["p"] = out["p"].astype(float)
    raw = out["p"].values
    _, fdr_p, _, _ = multipletests(raw, method="fdr_bh")
    out["p_fdr_bh"] = fdr_p
    out["raw_significance"] = out["p"].apply(significance_stars)
    out["fdr_significance"] = out["p_fdr_bh"].apply(significance_stars)
    out["p_fdr_bh"] = out["p_fdr_bh"].apply(lambda x: f"{float(x):.2e}")
    out["p"] = out["p"].apply(lambda x: f"{float(x):.2e}")
    out = out.drop(columns=["significance", "test_statistic", "-log2(p)"], errors="ignore")

    adj_cols = ("p_fdr_bh", "raw_significance", "fdr_significance")
    cols = [c for c in out.columns if c not in adj_cols]
    if "p" in cols:
        insert_at = cols.index("p") + 1
        for c in adj_cols:
            if c in out.columns:
                cols.insert(insert_at, c)
                insert_at += 1
        out = out[cols]
    return out


def create_pair_km_plot(
    data: pd.DataFrame,
    left_group: str,
    right_group: str,
    data_column: str,
    colors: Dict[str, str],
    style: KMPlotStyle,
    output_path: Path,
    *,
    color_overrides: Optional[Dict[Tuple[str, str], str]] = None,
) -> Optional[Dict[str, Any]]:
    """Create a two-group KM figure and return one pairwise summary row."""
    pair_groups = [left_group, right_group]
    pair_data = data[data["GROUP"].isin(pair_groups)].copy()
    if pair_data.empty:
        LOGGER.warning("No data for pair %s vs %s", left_group, right_group)
        return None

    style.apply_rcparams()
    pair_fig, pair_ax = plt.subplots(figsize=(5, 5.5))
    kmf_pair_list: List[KaplanMeierFitter] = []

    for group in pair_groups:
        group_data = pair_data[pair_data["GROUP"] == group].copy()
        if group_data.empty:
            continue
        group_data.loc[:, data_column] = pd.to_numeric(group_data[data_column], errors="coerce")
        group_data = group_data.dropna(subset=[data_column])
        kmf = KaplanMeierFitter()
        kmf.fit(group_data[data_column], event_observed=group_data["CENSOR"], label=group)
        plot_color = colors.get(group)
        if color_overrides and (left_group, right_group) in color_overrides and group == left_group:
            plot_color = color_overrides[(left_group, right_group)]
        kmf.plot(ax=pair_ax, color=plot_color, ci_show=style.cishow, linestyle="-", linewidth=1)
        kmf_pair_list.append(kmf)

    if not kmf_pair_list:
        plt.close(pair_fig)
        return None

    left_data = pair_data[pair_data["GROUP"] == left_group].dropna(subset=[data_column])
    right_data = pair_data[pair_data["GROUP"] == right_group].dropna(subset=[data_column])

    def _safe_median(value: Any) -> str:
        if pd.isna(value) or np.isinf(value):
            return "NR"
        return f"{float(value):.1f}"

    left_median = KaplanMeierFitter().fit(
        left_data[data_column], event_observed=left_data["CENSOR"]
    ).median_survival_time_
    right_median = KaplanMeierFitter().fit(
        right_data[data_column], event_observed=right_data["CENSOR"]
    ).median_survival_time_

    logrank_res = logrank_test(
        left_data[data_column],
        right_data[data_column],
        event_observed_A=left_data["CENSOR"],
        event_observed_B=right_data["CENSOR"],
    )

    cox_df = pair_data[[data_column, "CENSOR", "GROUP"]].dropna()
    cox_df["group_indicator"] = (cox_df["GROUP"] == left_group).astype(int)
    hr = ci_low = ci_high = hr_p = np.nan
    hr_text = "HR = NA"
    if cox_df["group_indicator"].nunique() == 2:
        cph = CoxPHFitter()
        cph.fit(cox_df[[data_column, "CENSOR", "group_indicator"]], duration_col=data_column, event_col="CENSOR")
        hr = float(np.exp(cph.params_["group_indicator"]))
        ci_low = float(np.exp(cph.confidence_intervals_.loc["group_indicator", "95% lower-bound"]))
        ci_high = float(np.exp(cph.confidence_intervals_.loc["group_indicator", "95% upper-bound"]))
        hr_p = float(cph.summary.loc["group_indicator", "p"])
        hr_text = f"HR = {hr:.2f} ({ci_low:.2f}-{ci_high:.2f})\n$p$ = {hr_p:.2e}"

    apply_km_axis_style(pair_ax, style)
    pair_ax.legend(labels=pair_groups, frameon=False, fontsize=style.font_legend, loc="upper right", bbox_to_anchor=(0.98, 0.98))
    pair_ax.text(
        0.52,
        0.40,
        space_plot_operators(
            f"{left_group} mPFS = {_safe_median(left_median)}\n"
            f"{right_group} mPFS = {_safe_median(right_median)}\n"
            f"logrank $p$ = {logrank_res.p_value:.2e}"
        ),
        transform=pair_ax.transAxes,
        fontsize=style.font * 1.1,
        va="top",
        bbox=dict(facecolor="white", alpha=0.85, edgecolor="none", pad=2),
    )
    pair_ax.text(
        0.52,
        0.26,
        space_plot_operators(hr_text),
        transform=pair_ax.transAxes,
        fontsize=style.font * 1.1,
        va="top",
        linespacing=1.0,
        bbox=dict(facecolor="white", alpha=0.85, edgecolor="none", pad=2),
    )
    add_at_risk_counts(
        *kmf_pair_list,
        ypos=-0.35,
        ax=pair_ax,
        rows_to_show=["At risk"],
        at_risk_count_from_start_of_period=True,
        xticks=KM_XTICKS,
    )
    plt.subplots_adjust(bottom=0.25, top=0.9, hspace=0.5)
    plt.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pair_fig.savefig(output_path, dpi=300)
    plt.close(pair_fig)

    return {
        "comparison": f"{left_group} vs {right_group}",
        f"mPFS_{left_group}": left_median,
        f"mPFS_{right_group}": right_median,
        "logrank_p": logrank_res.p_value,
        "HR": hr,
        "HR_CI_lower_95": ci_low,
        "HR_CI_upper_95": ci_high,
        "HR_P": hr_p,
    }


def export_global_pairwise_logrank(
    patient_data: pd.DataFrame,
    data_column: str,
    output_path: Path,
    *,
    robust: bool = False,
) -> pd.DataFrame:
    """Run all-pairs log-rank test and write colour-coded Excel export."""
    kwargs = {"specify_robust": True} if robust else {}
    results = pairwise_logrank_test(
        patient_data[data_column], patient_data["GROUP"], patient_data["CENSOR"], **kwargs
    )
    summary = format_pairwise_logrank_summary(results.summary)
    export_pairwise_logrank_excel(summary, str(output_path))
    return summary


# =============================================================================
# Analysis 1: On-label / off-label KM
# =============================================================================


def build_on_label_dataset(paths: Paths, source_path: Path) -> pd.DataFrame:
    """Construct grouped survival table for on-label analysis.

    Group definitions:
      Low:    LEVEL < 0
      Int:    0 <= LEVEL < 1000
      High:   LEVEL > 1000
      On:     LEVEL > 1000, On_label = IGAZ, APPROVED = IGAZ
      Off:    LEVEL > 1000, On_label = HAMIS, APPROVED = IGAZ
      Off+Ex: LEVEL > 1000, On_label = HAMIS (all approval statuses)
      SC:     standard-of-care chemotherapy reference cohort
    """
    patient = pd.read_excel(source_path)
    patient = harmonize_column_names(patient)
    patient = normalize_id_columns(patient)
    patient = clean_base_patient_table(patient)
    patient["CENSOR"] = 1
    validate_on_label_source_columns(patient)
    on_col, on_yes, on_no = resolve_on_label_columns(patient)
    chemo = load_chemo_from_xlsx(paths)

    result_first = patient[patient["LEVEL"] < 0].copy()
    result_second = patient[(patient["LEVEL"] >= 0) & (patient["LEVEL"] < 1000)].copy()
    result_third = patient[patient["LEVEL"] > 1000].copy()
    result_fourth = patient[
        (patient["LEVEL"] > 1000) & (patient[on_col] == on_yes) & (patient["APPROVED"] == "IGAZ")
    ].copy()
    result_fifth = patient[
        (patient["LEVEL"] > 1000) & (patient[on_col] == on_no) & (patient["APPROVED"] == "IGAZ")
    ].copy()
    result_sixth = patient[(patient["LEVEL"] > 1000) & (patient[on_col] == on_no)].copy()

    result_first["GROUP"] = "Low"
    result_second["GROUP"] = "Int"
    result_third["GROUP"] = "High"
    result_fourth["GROUP"] = "On"
    result_fifth["GROUP"] = "Off"
    result_sixth["GROUP"] = "Off+Ex"

    merged_groups = pd.concat(
        [result_first, result_second, result_third, result_fourth, result_fifth, result_sixth]
    )
    merged_groups["GROUP"] = merged_groups["GROUP"].replace({"LOW": "Low", "low": "Low"})
    patient_out = pd.concat([merged_groups, chemo], ignore_index=True)
    patient_out["CENSOR"] = 1
    return patient_out


def run_on_label_analysis(
    paths: Paths, spec: KMAnalysisSpec, output_dir: Path, source_path: Path
) -> None:
    LOGGER.info("Starting on-label KM analysis -> %s", output_dir)
    patient_data = build_on_label_dataset(paths, source_path)
    style = KMPlotStyle()
    colors = {
        "High": "#2E8B57",
        "On": "#50C878",
        "Off": "#0F52BA",
        "Off+Ex": "#0f7eba",
        "Int": "#F0C864",
        "SC": "#b121f6",
        "Low": "#f71c6c",
    }
    legend_labels = ["Low", "Int", "High", "High On", "High Off", "High Off+Exp", "SC"]

    fig, _, _, summary_df = fit_multigroup_km_plot(
        patient_data,
        spec.data_column,
        colors,
        style,
        legend_labels=legend_labels,
        at_risk_ypos=-0.4,
    )
    primary = spec.primary_tumor
    fig.savefig(output_dir / f"{primary}_Kmplot_{legacy_output_folder_name(spec)}_{spec.date}.jpg", dpi=300)
    plt.close(fig)

    summary_df.to_excel(output_dir / f"{primary}_median_survival_times.xlsx", index=False)

    pairwise_rows: List[Dict[str, Any]] = []
    for left, right in spec.pair_comparisons:
        pair_name = f"{left}_vs_{right}".replace("+", "plus").replace(" ", "")
        row = create_pair_km_plot(
            patient_data,
            left,
            right,
            spec.data_column,
            colors,
            style,
            output_dir / f"{primary}_Kmplot_{pair_name}_{legacy_output_folder_name(spec)}_{spec.date}.jpg",
        )
        if row:
            pairwise_rows.append(row)

    pairwise_hr_df = append_fdr_to_pairwise_hr(pd.DataFrame(pairwise_rows))
    export_pairwise_logrank_excel(
        pairwise_hr_df,
        str(output_dir / f"{primary}_pairwise_HR_values_{legacy_output_folder_name(spec)}_{spec.date}.xlsx"),
    )
    export_global_pairwise_logrank(
        patient_data,
        spec.data_column,
        output_dir / f"{primary}_pairwise_logrank_{legacy_output_folder_name(spec)}_{spec.date}_.xlsx",
        robust=True,
    )
    LOGGER.info("Completed on-label KM analysis.")


# =============================================================================
# Analysis 2: Per-patient ranking KM
# =============================================================================


def build_ranking_dataset(
    paths: Paths, source_path: Path
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Assign dense rank groups per patient and append SC reference rows.

    Returns (data_without_sc, data_with_sc) for main and pairwise plots.
    """
    patient = pd.read_excel(source_path)
    patient = harmonize_column_names(patient)
    patient = normalize_id_columns(patient)
    patient = clean_base_patient_table(patient)
    patient["CENSOR"] = True

    chemo = load_chemo_from_csv(paths)
    ranked = assign_dense_level_groups(patient, model_col="Model", level_col="LEVEL", group_col="GROUP")
    combined = pd.concat([ranked, chemo], ignore_index=True)

    with_sc = normalize_survival_columns(combined.copy(), DEFAULT_DATA_COLUMN)
    with_sc = with_sc.dropna(subset=[DEFAULT_DATA_COLUMN])

    without_sc = with_sc[with_sc["GROUP"] != "SC"].copy()
    return without_sc, with_sc


def run_ranking_analysis(
    paths: Paths, spec: KMAnalysisSpec, output_dir: Path, source_path: Path
) -> None:
    LOGGER.info("Starting ranking KM analysis -> %s", output_dir)
    patient_data, patient_data_with_sc = build_ranking_dataset(paths, source_path)
    style = KMPlotStyle()
    colors = {
        "1": "#FFD700",
        "2": "#C0C0C0",
        "3": "#CD7F32",
        "4": "#C2185B",
        "5": "#64B5F6",
        "6": "#D8F3DC",
        "7-10": "#E8D5FF",
        "SC": "#b121f6",
    }

    fig, _, _, summary_df = fit_multigroup_km_plot(
        patient_data,
        spec.data_column,
        colors,
        style,
        legend_title="Ranking",
        at_risk_ypos=-0.3,
    )
    primary = safe_filename(spec.primary_tumor) or "all"
    fig.savefig(
        output_dir / f"{primary}_Kmplot_{spec.legacy_name}_{spec.data_column}_{spec.date}.jpg",
        dpi=300,
    )
    plt.close(fig)
    summary_df.to_excel(
        output_dir / f"{primary}_median_survival_times_{legacy_output_folder_name(spec)}.xlsx",
        index=False,
    )

    pairwise_rows: List[Dict[str, Any]] = []
    for rank_group in spec.extra.get("ranking_groups_vs_sc", ()):
        row = create_pair_km_plot(
            patient_data_with_sc,
            rank_group,
            "SC",
            spec.data_column,
            colors,
            style,
            output_dir / f"{primary}_Kmplot_{rank_group}_vs_SC_{legacy_output_folder_name(spec)}_{spec.date}.jpg",
            color_overrides={(rank_group, "SC"): "#404040"} if rank_group == "7-10" else None,
        )
        if row:
            pairwise_rows.append(row)

    pairwise_hr_df = append_fdr_to_pairwise_hr(pd.DataFrame(pairwise_rows))
    export_pairwise_logrank_excel(
        pairwise_hr_df,
        str(output_dir / f"{primary}_pairwise_HR_values_{legacy_output_folder_name(spec)}_{spec.date}.xlsx"),
    )
    export_global_pairwise_logrank(
        patient_data,
        spec.data_column,
        output_dir / f"{primary}_pairwise_logrank_{legacy_output_folder_name(spec)}_{spec.date}_.xlsx",
    )
    LOGGER.info("Completed ranking KM analysis.")


# =============================================================================
# Analysis 3: SHIVA-tier KM
# =============================================================================


def build_shiva_dataset(
    paths: Paths, source_path: Path
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Build Low/Int/High + All + SC + Non_assigned survival table.

    SHIVA tier thresholds: Low < 0; 0 <= Int < 1000; High >= 1000.
    Returns (combined_all_groups, subset_without_sc_all).
    """
    if source_path.suffix.lower() == ".csv":
        patient_csv = pd.read_csv(source_path)
    else:
        patient_csv = pd.read_excel(source_path)

    patient_csv = harmonize_column_names(patient_csv)
    patient_csv = normalize_id_columns(patient_csv)
    patient_csv = clean_base_patient_table(patient_csv)
    patient_csv["CENSOR"] = True

    chemo = load_chemo_from_csv(paths)
    non_assigned = load_non_assigned_table(paths)

    low = patient_csv[patient_csv["LEVEL"] < 0].copy()
    inter = patient_csv[(patient_csv["LEVEL"] >= 0) & (patient_csv["LEVEL"] < 1000)].copy()
    high = patient_csv[patient_csv["LEVEL"] > 1000].copy()
    low["GROUP"], inter["GROUP"], high["GROUP"] = "Low", "Int", "High"

    tiered = pd.concat([low, inter, high], ignore_index=True)
    all_rows = patient_csv.copy()
    all_rows["GROUP"] = "All"
    combined = pd.concat([all_rows, tiered, chemo], ignore_index=True)
    subset = pd.concat([tiered.copy(), non_assigned], ignore_index=True)

    combined = normalize_survival_columns(combined, DEFAULT_DATA_COLUMN)
    subset = normalize_survival_columns(subset, DEFAULT_DATA_COLUMN)
    return combined, subset


def run_shiva_analysis(
    paths: Paths, spec: KMAnalysisSpec, output_dir: Path, source_path: Path
) -> None:
    LOGGER.info("Starting SHIVA KM analysis -> %s", output_dir)
    patient_data, subset_data = build_shiva_dataset(paths, source_path)
    style = KMPlotStyle()
    colors = {
        "High": "#50C878",
        "Int": "#ffbc15",
        "Low": "#f71c6c",
        "SC": "#b121f6",
        "All": "#000000",
        "Non_assigned": "#8B4513",
    }
    subset_order = ["High", "Int", "Low", "Non_assigned"]

    fig, _, _, summary_df = fit_multigroup_km_plot(
        patient_data,
        spec.data_column,
        colors,
        style,
        label_suffix_n=True,
        at_risk_ypos=-0.3,
    )
    summary_df["Median Survival Time"] = summary_df["Median Survival Time"].round(1)
    summary_df["CI_lower"] = summary_df["CI_lower"].round(1)
    summary_df["CI_upper"] = summary_df["CI_upper"].round(1)
    fig.tight_layout(rect=[0.0, 0.12, 1.0, 1.0])
    fig.savefig(
        output_dir / f"_Kmplot_{legacy_output_folder_name(spec)}_{spec.date}_with_with_CI.jpg",
        dpi=300,
        bbox_inches="tight",
    )
    plt.close(fig)

    fig2, _, _, _ = fit_multigroup_km_plot(
        subset_data,
        spec.data_column,
        colors,
        style,
        group_order=subset_order,
        label_suffix_n=True,
        legend_title="Tiers",
        at_risk_ypos=-0.3,
    )
    fig2.tight_layout(rect=[0.0, 0.12, 1.0, 1.0])
    fig2.savefig(
        output_dir / f"_Kmplot_{legacy_output_folder_name(spec)}_{spec.date}_High_Int_Low_Non_assigned.jpg",
        dpi=300,
        bbox_inches="tight",
    )
    plt.close(fig2)

    summary_df.to_excel(output_dir / "_median_survival_times.xlsx", index=False)
    export_global_pairwise_logrank(
        patient_data,
        spec.data_column,
        output_dir / f"_pairwise_logrank_{legacy_output_folder_name(spec)}_{spec.date}_.xlsx",
    )
    LOGGER.info("Completed SHIVA KM analysis.")


# =============================================================================
# Orchestration
# =============================================================================


def build_analysis_specs() -> List[KMAnalysisSpec]:
    return [
        KMAnalysisSpec(
            run_key="on_label",
            source_script_stem="KM_ON_LABEL_final",
            legacy_name="KM_SHIVA_onLabel_offlabel_chemo_combined",
            date="2026_O6_12_",
            pair_comparisons=(
                ("Low", "SC"),
                ("Low", "Int"),
                ("Low", "High"),
                ("Int", "SC"),
                ("High", "SC"),
            ),
        ),
        KMAnalysisSpec(
            run_key="ranking",
            source_script_stem="KM_ranking_final_OI_Jun2O26",
            legacy_name="KM_primary_ranking",
            date="2026_06_15_",
            extra={"ranking_groups_vs_sc": ("1", "2", "3", "4", "5", "6", "7-10")},
        ),
        KMAnalysisSpec(
            run_key="shiva",
            source_script_stem="KM_SHIVA_final",
            legacy_name="SHIVA__SC_KM_Kaplan_Meier",
            date="2026_O6_12_",
            primary_tumor="skin",
        ),
    ]


def resolve_paths(args: argparse.Namespace) -> Paths:
    source_primary = args.source.resolve() if args.source else DEFAULT_SOURCE_XLSX.resolve()
    return Paths(
        project_root=REPO_ROOT,
        script_dir=_SCRIPT_DIR,
        source_primary_xlsx=source_primary,
        chemo_csv=DATA_DIR / "chemo_data_2024_12_06.csv",
        non_assigned_xlsx=DATA_DIR / "non_assigned_merged.xlsx",
    )


def source_for_analysis(paths: Paths, _spec: KMAnalysisSpec, _args: argparse.Namespace) -> Path:
    """All analyses use the same primary workbook."""
    return paths.source_primary_xlsx


def write_analysis_info(output_dir: Path, spec: KMAnalysisSpec, source_path: Path) -> None:
    info = {
        "run_key": spec.run_key,
        "source_script_stem": spec.source_script_stem,
        "legacy_name": spec.legacy_name,
        "date": spec.date,
        "data_column": spec.data_column,
        "legacy_output_folder": legacy_output_folder_name(spec),
        "source_path": str(source_path),
        "source_meta": file_meta(source_path),
    }
    (output_dir.parent / "analysis_info.json").write_text(json.dumps(info, indent=2), encoding="utf-8")


def main() -> None:
    """Run requested KM analyses and write outputs with reproducibility metadata."""
    args = parse_args()
    paths = resolve_paths(args)
    output_root = args.output_dir.resolve() / KM_PDX_OUTPUT_DIRNAME
    output_root.mkdir(parents=True, exist_ok=True)

    if not paths.source_primary_xlsx.exists():
        raise FileNotFoundError(
            f"Primary source not found: {paths.source_primary_xlsx}\n"
            f"Expected: data/{DEFAULT_SOURCE_XLSX.name}"
        )

    source_meta = validate_configured_sources(paths)
    LOGGER.info("All analyses using primary source: %s", paths.source_primary_xlsx)
    selected = "all" if "all" in args.analyses else ",".join(args.analyses)

    runners: Dict[str, Callable[..., None]] = {
        "on_label": run_on_label_analysis,
        "ranking": run_ranking_analysis,
        "shiva": run_shiva_analysis,
    }

    analysis_folders: Dict[str, str] = {}
    for spec in build_analysis_specs():
        if spec.run_key not in runners:
            continue
        if "all" not in args.analyses and spec.run_key not in args.analyses:
            LOGGER.info("Skipping analysis '%s' (not in --analyses).", spec.run_key)
            continue

        source_path = source_for_analysis(paths, spec, args)
        out_dir = resolve_analysis_output_dir(output_root, spec)
        analysis_folders[spec.run_key] = str(out_dir)
        write_analysis_info(out_dir, spec, source_path)
        runners[spec.run_key](paths, spec, out_dir, source_path)

    metadata = {
        "seed": SEED,
        "generated_at": datetime.now().isoformat(),
        "output_root": str(output_root),
        "analyses_requested": selected,
        "analysis_output_folders": analysis_folders,
        **source_meta,
    }
    (output_root / "run_metadata.json").write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    write_output_pointer(_SCRIPT_DIR, output_root)
    LOGGER.info("KM_PDX completed. Output: %s", output_root)


if __name__ == "__main__":
    main()
